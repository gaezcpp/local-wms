# -*- coding: utf-8 -*-
import base64
from io import BytesIO

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except Exception:
    openpyxl = None


class TaggingBOMImportWizard(models.TransientModel):
    _name = "tagging.bom.import.wizard"
    _description = "Import Equipment Tree / BOM from Excel"

    file = fields.Binary(string="File Excel", required=True)
    filename = fields.Char(string="Filename")

    def action_import(self):
        self.ensure_one()

        if not openpyxl:
            raise UserError(_("Library openpyxl belum tersedia di server Odoo."))

        if not self.file:
            raise UserError(_("Mohon upload file Excel."))

        data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        ws = wb.active  # ambil sheet pertama

        # Ambil header
        headers = []
        for cell in ws[1]:
            headers.append((cell.value or "").strip() if isinstance(cell.value, str) else cell.value)

        def norm(h):
            return (h or "").strip().lower()

        header_map = {norm(h): idx for idx, h in enumerate(headers)}

        required = [
            "sistem",
            "sub_sistem",
            "unit_mesin",
            "bagian_mesin",
            "spare_part",
            "spesifikasi_spare_part",
            "sku",
            "bu",
        ]
        missing = [h for h in required if h not in header_map]
        if missing:
            raise UserError(_("Header Excel tidak lengkap. Kurang: %s") % ", ".join(missing))

        System = self.env["tagging.system"].sudo()
        SubSystem = self.env["tagging.subsystem"].sudo()
        Unit = self.env["tagging.machine_unit"].sudo()
        Part = self.env["tagging.machine_part"].sudo()
        Spare = self.env["tagging.spare_part"].sudo()
        BOM = self.env["tagging.machine_bom"].sudo()
        BU = self.env["tagging.bu"].sudo()

        # Cache biar cepat (minim query)
        cache_system = {}
        cache_subsystem = {}
        cache_unit = {}
        cache_part = {}
        cache_spare = {}
        cache_bu = {}

        def clean(v):
            if v is None:
                return ""
            if isinstance(v, str):
                return v.strip()
            return str(v).strip()

        # ---------- GET/CREATE HELPERS (SEMUA TOLERAN KOSONG) ----------

        def get_bu(bu_val):
            """BU wajib untuk unit_mesin (bu_id required). Kalau kosong -> return False supaya bisa di-skip."""
            key = clean(bu_val)
            if not key:
                return False
            if key in cache_bu:
                return cache_bu[key]

            rec = BU.search([("name", "=", key)], limit=1)
            if not rec and "code" in BU._fields:
                rec = BU.search([("code", "=", key)], limit=1)

            if not rec:
            # auto-create BU biar import lanjut
                vals = {"name": key}
                if "code" in BU._fields:
                    vals["code"] = key
                rec = BU.create(vals)


            cache_bu[key] = rec.id
            return rec.id

        def get_or_create_system(name):
            name = clean(name)
            if not name:
                return False
            if name in cache_system:
                return cache_system[name]
            rec = System.search([("name", "=", name)], limit=1)
            if not rec:
                rec = System.create({"name": name})
            cache_system[name] = rec.id
            return rec.id

        def get_or_create_subsystem(name, system_id):
            name = clean(name)
            if not name or not system_id:
                return False
            key = (system_id, name)
            if key in cache_subsystem:
                return cache_subsystem[key]
            rec = SubSystem.search([("name", "=", name), ("system_id", "=", system_id)], limit=1)
            if not rec:
                rec = SubSystem.create({"name": name, "system_id": system_id})
            cache_subsystem[key] = rec.id
            return rec.id

        def get_or_create_unit(name, subsystem_id, bu_id):
            name = clean(name)
            if not name or not subsystem_id or not bu_id:
                return False
            key = (subsystem_id, bu_id, name)
            if key in cache_unit:
                return cache_unit[key]

            domain = [("name", "=", name), ("subsystem_id", "=", subsystem_id), ("bu_id", "=", bu_id)]
            rec = Unit.search(domain, limit=1)
            if not rec:
                rec = Unit.create({"name": name, "subsystem_id": subsystem_id, "bu_id": bu_id})
            cache_unit[key] = rec.id
            return rec.id

        def get_or_create_part(name, unit_id):
            name = clean(name)
            if not name or not unit_id:
                return False
            key = (unit_id, name)
            if key in cache_part:
                return cache_part[key]
            rec = Part.search([("name", "=", name), ("unit_id", "=", unit_id)], limit=1)
            if not rec:
                rec = Part.create({"name": name, "unit_id": unit_id})
            cache_part[key] = rec.id
            return rec.id

        def get_or_create_spare(name, spec, sku, bu_id):
            name = clean(name)
            if not name:
                return False

            spec = clean(spec)
            sku = clean(sku)

            # Kunci cache berdasarkan (name, sku, bu) biar tidak nyampur
            key = (name, sku, bu_id)
            if key in cache_spare:
                return cache_spare[key]

            domain = [("name", "=", name)]
            if sku:
                domain.append(("sku", "=", sku))
            if bu_id:
                domain.append(("bu_id", "=", bu_id))

            rec = Spare.search(domain, limit=1)
            if not rec:
                rec = Spare.create(
                    {
                        "name": name,
                        "specification": spec or False,
                        "sku": sku or False,
                        "bu_id": bu_id or False,
                    }
                )
            else:
                # Optional: update spec/sku/bu kalau kosong di master
                vals = {}
                if spec and not rec.specification:
                    vals["specification"] = spec
                if sku and not rec.sku:
                    vals["sku"] = sku
                if bu_id and not rec.bu_id:
                    vals["bu_id"] = bu_id
                if vals:
                    rec.write(vals)

            cache_spare[key] = rec.id
            return rec.id

        # ---------- COUNTERS ----------
        created = 0
        updated = 0

        skipped_bu = 0
        skipped_system = 0
        skipped_subsystem = 0
        skipped_unit = 0
        skipped_part = 0
        skipped_spare = 0

        # Loop rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            sistem = row[header_map["sistem"]]
            sub_sistem = row[header_map["sub_sistem"]]
            unit_mesin = row[header_map["unit_mesin"]]
            bagian_mesin = row[header_map["bagian_mesin"]]
            spare_part = row[header_map["spare_part"]]
            spec = row[header_map["spesifikasi_spare_part"]]
            sku = row[header_map["sku"]]
            bu = row[header_map["bu"]]

            # BU wajib (karena tagging.machine_unit.bu_id required)
            bu_id = get_bu(bu)
            if not bu_id:
                skipped_bu += 1
                continue

            system_id = get_or_create_system(sistem)
            if not system_id:
                skipped_system += 1
                continue

            subsystem_id = get_or_create_subsystem(sub_sistem, system_id)
            if not subsystem_id:
                skipped_subsystem += 1
                continue

            unit_id = get_or_create_unit(unit_mesin, subsystem_id, bu_id)
            if not unit_id:
                skipped_unit += 1
                continue

            # part wajib di BOM (required=True)
            part_id = get_or_create_part(bagian_mesin, unit_id)
            if not part_id:
                skipped_part += 1
                continue

            # spare wajib di BOM (required=True)
            spare_id = get_or_create_spare(spare_part, spec, sku, bu_id)
            if not spare_id:
                skipped_spare += 1
                continue

            # BOM uniq by kombinasi relasi
            dom = [
                ("system_id", "=", system_id),
                ("subsystem_id", "=", subsystem_id),
                ("unit_id", "=", unit_id),
                ("part_id", "=", part_id),
                ("spare_part_id", "=", spare_id),
            ]
            bom = BOM.search(dom, limit=1)

            vals = {
                "system_id": system_id,
                "subsystem_id": subsystem_id,
                "unit_id": unit_id,
                "part_id": part_id,
                "spare_part_id": spare_id,
                # snapshot dari excel
                "specification": clean(spec) or False,
                "sku": clean(sku) or False,
                "bu_id": bu_id or False,
                "active": True,
            }

            if bom:
                bom.write(vals)
                updated += 1
            else:
                BOM.create(vals)
                created += 1

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import Selesai"),
                "message": _(
                    "created:%s | updated:%s | skip_bu:%s | skip_system:%s | skip_subsystem:%s | "
                    "skip_unit:%s | skip_part:%s | skip_spare:%s"
                )
                % (
                    created,
                    updated,
                    skipped_bu,
                    skipped_system,
                    skipped_subsystem,
                    skipped_unit,
                    skipped_part,
                    skipped_spare,
                ),
                "type": "success",
                "sticky": False,
            },
        }
